"""
Professional Excel portfolio report generation service.

This module converts an existing PortfolioReport into a professional XLSX
workbook.

Responsibilities
----------------
- Validate Excel generation inputs.
- Build styled portfolio reporting worksheets.
- Render report metadata.
- Render portfolio performance metrics.
- Render historical analytics when available.
- Render advanced-analytics availability information.
- Render AI insights, notes, warnings, and disclaimer content.
- Return XLSX bytes or save them to a caller-provided path.
- Prepare workbook data for Streamlit download buttons.

This module performs no portfolio calculations and does not retrieve data
from portfolio or analytics services. It consumes PortfolioReport only.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import date, datetime, timezone
from io import BytesIO
from math import isfinite
from pathlib import Path
from typing import Any, Final

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from services.reporting.report_models import PortfolioReport


# ============================================================
# Constants
# ============================================================

DEFAULT_EXCEL_FILENAME: Final[str] = "portfolio_report.xlsx"

EXCEL_MIME_TYPE: Final[str] = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)

UNAVAILABLE_VALUE: Final[str] = "Unavailable"

MAX_CELL_TEXT_LENGTH: Final[int] = 32_000

TITLE_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    fgColor="17365D",
)

SECTION_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

HEADER_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    fgColor="4472C4",
)

LABEL_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    fgColor="EEF3F8",
)

WARNING_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)

THIN_BORDER: Final[Border] = Border(
    left=Side(
        style="thin",
        color="D9E1EA",
    ),
    right=Side(
        style="thin",
        color="D9E1EA",
    ),
    top=Side(
        style="thin",
        color="D9E1EA",
    ),
    bottom=Side(
        style="thin",
        color="D9E1EA",
    ),
)

TITLE_FONT: Final[Font] = Font(
    name="Calibri",
    size=18,
    bold=True,
    color="FFFFFF",
)

SECTION_FONT: Final[Font] = Font(
    name="Calibri",
    size=12,
    bold=True,
    color="17365D",
)

HEADER_FONT: Final[Font] = Font(
    name="Calibri",
    size=11,
    bold=True,
    color="FFFFFF",
)

LABEL_FONT: Final[Font] = Font(
    name="Calibri",
    size=10,
    bold=True,
    color="17365D",
)

VALUE_FONT: Final[Font] = Font(
    name="Calibri",
    size=10,
    color="25313C",
)

SMALL_FONT: Final[Font] = Font(
    name="Calibri",
    size=9,
    color="5A6573",
)

WARNING_FONT: Final[Font] = Font(
    name="Calibri",
    size=10,
    color="7C2D12",
)

CURRENCY_NUMBER_FORMAT: Final[str] = (
    '₹#,##0.00;[Red]-₹#,##0.00'
)

PERCENTAGE_NUMBER_FORMAT: Final[str] = (
    '0.00%;[Red]-0.00%'
)

DATE_NUMBER_FORMAT: Final[str] = "dd mmm yyyy"

DATETIME_NUMBER_FORMAT: Final[str] = (
    "dd mmm yyyy hh:mm"
)


# ============================================================
# Exceptions
# ============================================================


class ExcelReportError(RuntimeError):
    """
    Base exception raised by Excel report generation.
    """


class ExcelReportValidationError(
    ExcelReportError
):
    """
    Raised when Excel report input is invalid.
    """


class ExcelReportGenerationError(
    ExcelReportError
):
    """
    Raised when Excel workbook generation or saving fails.
    """


# ============================================================
# Validation
# ============================================================


def _validate_report(
    report: PortfolioReport,
) -> PortfolioReport:
    """
    Validate the report supplied to the Excel service.

    Args:
        report:
            Expected PortfolioReport instance.

    Returns:
        Validated PortfolioReport.

    Raises:
        TypeError:
            When report is not a PortfolioReport.
    """

    if not isinstance(
        report,
        PortfolioReport,
    ):
        raise TypeError(
            "report must be a PortfolioReport."
        )

    return report


def _validate_output_path(
    output_path: str | Path,
) -> Path:
    """
    Validate and normalize an Excel output path.

    Args:
        output_path:
            Target path for the generated workbook.

    Returns:
        Resolved absolute path ending in ``.xlsx``.

    Raises:
        TypeError:
            When output_path is not a string or Path.

        ExcelReportValidationError:
            When the path is blank or does not use an XLSX extension.
    """

    if not isinstance(
        output_path,
        (str, Path),
    ):
        raise TypeError(
            "output_path must be a string or Path."
        )

    if (
        isinstance(
            output_path,
            str,
        )
        and not output_path.strip()
    ):
        raise ExcelReportValidationError(
            "output_path cannot be blank."
        )

    path = Path(
        output_path
    ).expanduser()

    if path.suffix.lower() != ".xlsx":
        raise ExcelReportValidationError(
            "output_path must end with .xlsx."
        )

    return path.resolve()


def _validate_filename(
    filename: str,
) -> str:
    """
    Validate a downloadable Excel filename.

    Args:
        filename:
            Download filename without directory components.

    Returns:
        Normalized XLSX filename.

    Raises:
        TypeError:
            When filename is not a string.

        ExcelReportValidationError:
            When filename is blank, contains directory components,
            or does not end with ``.xlsx``.
    """

    if not isinstance(
        filename,
        str,
    ):
        raise TypeError(
            "filename must be a string."
        )

    normalized = filename.strip()

    if not normalized:
        raise ExcelReportValidationError(
            "filename cannot be blank."
        )

    if (
        Path(normalized).name != normalized
        or "/" in normalized
        or "\\" in normalized
    ):
        raise ExcelReportValidationError(
            "filename must not contain directory components."
        )

    if not normalized.lower().endswith(
        ".xlsx"
    ):
        raise ExcelReportValidationError(
            "filename must end with .xlsx."
        )

    return normalized


def _validate_boolean(
    value: bool,
    *,
    parameter_name: str,
) -> bool:
    """
    Validate a boolean service option.

    Args:
        value:
            Value to validate.

        parameter_name:
            Name included in the validation error message.

    Returns:
        Validated boolean.

    Raises:
        TypeError:
            When value is not a boolean or parameter_name is not a string.

        ExcelReportValidationError:
            When parameter_name is blank.
    """

    if not isinstance(
        parameter_name,
        str,
    ):
        raise TypeError(
            "parameter_name must be a string."
        )

    normalized_name = parameter_name.strip()

    if not normalized_name:
        raise ExcelReportValidationError(
            "parameter_name cannot be blank."
        )

    if not isinstance(
        value,
        bool,
    ):
        raise TypeError(
            f"{normalized_name} must be a boolean."
        )

    return value


# ============================================================
# Value Formatting
# ============================================================


def _safe_cell_text(
    value: Any,
    *,
    maximum_length: int = MAX_CELL_TEXT_LENGTH,
) -> str:
    """
    Convert arbitrary report content into bounded worksheet text.

    Args:
        value:
            Value to convert.

        maximum_length:
            Maximum number of characters to retain.

    Returns:
        Normalized display text.
    """

    if isinstance(
        maximum_length,
        bool,
    ) or not isinstance(
        maximum_length,
        int,
    ):
        raise TypeError(
            "maximum_length must be an integer."
        )

    if maximum_length <= 0:
        raise ExcelReportValidationError(
            "maximum_length must be greater than zero."
        )

    if value is None:
        return UNAVAILABLE_VALUE

    if isinstance(
        value,
        bool,
    ):
        text = (
            "Yes"
            if value
            else "No"
        )

    elif isinstance(
        value,
        float,
    ):
        text = (
            f"{value:,.4f}"
            if isfinite(value)
            else UNAVAILABLE_VALUE
        )

    elif isinstance(
        value,
        Mapping,
    ):
        text = "; ".join(
            (
                f"{_safe_cell_text(key, maximum_length=500)}: "
                f"{_safe_cell_text(item, maximum_length=1_000)}"
            )
            for key, item in value.items()
        )

    elif isinstance(
        value,
        Sequence,
    ) and not isinstance(
        value,
        (str, bytes, bytearray),
    ):
        text = "; ".join(
            _safe_cell_text(
                item,
                maximum_length=1_000,
            )
            for item in value
        )

    else:
        text = str(value)

    normalized = " ".join(
        text.split()
    )

    if len(normalized) <= maximum_length:
        return normalized

    if maximum_length <= 3:
        return normalized[
            :maximum_length
        ]

    return (
        normalized[
            : maximum_length - 3
        ].rstrip()
        + "..."
    )


def _normalise_cell_value(
    value: Any,
) -> Any:
    """
    Convert unsupported or unsafe values into Excel-compatible content.

    Numeric and date values are retained so native workbook formatting can
    be applied. Timezone-aware datetimes are converted to timezone-naive UTC
    datetimes because Excel does not support timezone information.

    Complex objects are converted into bounded display text.
    """

    if value is None:
        return UNAVAILABLE_VALUE

    if isinstance(
        value,
        bool,
    ):
        return value

    if isinstance(
        value,
        int,
    ) and not isinstance(
        value,
        bool,
    ):
        return value

    if isinstance(
        value,
        float,
    ):
        if isfinite(value):
            return value

        return UNAVAILABLE_VALUE

    # datetime must be checked before date because datetime is a subclass
    # of date.
    if isinstance(
        value,
        datetime,
    ):
        if (
            value.tzinfo is not None
            and value.utcoffset() is not None
        ):
            return (
                value.astimezone(
                    timezone.utc
                )
                .replace(
                    tzinfo=None
                )
            )

        return value

    if isinstance(
        value,
        date,
    ):
        return value

    if isinstance(
        value,
        str,
    ):
        return _safe_cell_text(
            value
        )

    return _safe_cell_text(
        value
    )


# ============================================================
# Workbook Style Helpers
# ============================================================


def _create_workbook(
    report: PortfolioReport,
) -> Workbook:
    """
    Create a workbook configured for portfolio reporting.
    """

    validated_report = _validate_report(
        report
    )

    workbook = Workbook()

    workbook.properties.title = (
        validated_report.metadata.title
    )

    workbook.properties.subject = (
        "Professional mutual fund portfolio analytics report"
    )

    workbook.properties.creator = (
        validated_report.metadata.application_name
    )

    workbook.properties.description = (
        "Generated portfolio performance, historical analytics, "
        "advanced analytics, AI insights, notes, and warnings."
    )

    return workbook


def _validate_row_number(
    row: int,
    *,
    parameter_name: str = "row",
) -> int:
    """
    Validate a one-based worksheet row number.
    """

    if isinstance(
        row,
        bool,
    ) or not isinstance(
        row,
        int,
    ):
        raise TypeError(
            f"{parameter_name} must be an integer."
        )

    if row <= 0:
        raise ExcelReportValidationError(
            f"{parameter_name} must be greater than zero."
        )

    return row


def _validate_column_number(
    column: int,
    *,
    parameter_name: str = "column",
) -> int:
    """
    Validate a one-based worksheet column number.
    """

    if isinstance(
        column,
        bool,
    ) or not isinstance(
        column,
        int,
    ):
        raise TypeError(
            f"{parameter_name} must be an integer."
        )

    if column <= 0:
        raise ExcelReportValidationError(
            f"{parameter_name} must be greater than zero."
        )

    return column


def _apply_title_style(
    worksheet: Worksheet,
    *,
    row: int,
    title: str,
    end_column: int = 2,
) -> None:
    """
    Apply a merged title row to a worksheet.
    """

    _validate_row_number(row)

    _validate_column_number(
        end_column,
        parameter_name="end_column",
    )

    if not isinstance(
        title,
        str,
    ):
        raise TypeError(
            "title must be a string."
        )

    normalized_title = title.strip()

    if not normalized_title:
        raise ExcelReportValidationError(
            "title cannot be blank."
        )

    worksheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=end_column,
    )

    cell = worksheet.cell(
        row=row,
        column=1,
        value=normalized_title,
    )

    cell.fill = TITLE_FILL
    cell.font = TITLE_FONT
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[
        row
    ].height = 30


def _apply_section_style(
    worksheet: Worksheet,
    *,
    row: int,
    title: str,
    end_column: int = 2,
) -> None:
    """
    Apply a merged section heading to a worksheet.
    """

    _validate_row_number(row)

    _validate_column_number(
        end_column,
        parameter_name="end_column",
    )

    if not isinstance(
        title,
        str,
    ):
        raise TypeError(
            "title must be a string."
        )

    normalized_title = title.strip()

    if not normalized_title:
        raise ExcelReportValidationError(
            "title cannot be blank."
        )

    worksheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=end_column,
    )

    cell = worksheet.cell(
        row=row,
        column=1,
        value=normalized_title,
    )

    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    worksheet.row_dimensions[
        row
    ].height = 22


def _apply_header_style(
    cell: Cell,
) -> None:
    """
    Apply reusable table-header styling.
    """

    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )


def _apply_label_style(
    cell: Cell,
) -> None:
    """
    Apply reusable key-value label styling.
    """

    cell.fill = LABEL_FILL
    cell.font = LABEL_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )


def _apply_value_style(
    cell: Cell,
) -> None:
    """
    Apply reusable key-value value styling.
    """

    cell.font = VALUE_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )


def _apply_currency_style(
    cell: Cell,
) -> None:
    """
    Apply currency formatting to a numeric cell.
    """

    _apply_value_style(cell)

    cell.number_format = (
        CURRENCY_NUMBER_FORMAT
    )

    cell.alignment = Alignment(
        horizontal="right",
        vertical="top",
    )


def _apply_percentage_style(
    cell: Cell,
) -> None:
    """
    Apply percentage formatting to a decimal-fraction cell.
    """

    _apply_value_style(cell)

    cell.number_format = (
        PERCENTAGE_NUMBER_FORMAT
    )

    cell.alignment = Alignment(
        horizontal="right",
        vertical="top",
    )


def _apply_date_style(
    cell: Cell,
) -> None:
    """
    Apply date formatting to a cell.
    """

    _apply_value_style(cell)

    cell.number_format = (
        DATE_NUMBER_FORMAT
    )


def _apply_datetime_style(
    cell: Cell,
) -> None:
    """
    Apply datetime formatting to a cell.
    """

    _apply_value_style(cell)

    cell.number_format = (
        DATETIME_NUMBER_FORMAT
    )


def _write_key_value_table(
    worksheet: Worksheet,
    *,
    start_row: int,
    rows: Sequence[
        tuple[str, Any]
    ],
) -> int:
    """
    Write a styled two-column key-value table.

    Args:
        worksheet:
            Target worksheet.

        start_row:
            First output row.

        rows:
            Label-value pairs.

    Returns:
        Next available row following the table.
    """

    _validate_row_number(
        start_row,
        parameter_name="start_row",
    )

    if not isinstance(
        rows,
        Sequence,
    ) or isinstance(
        rows,
        (str, bytes, bytearray),
    ):
        raise TypeError(
            "rows must be a sequence."
        )

    current_row = start_row

    for item in rows:
        if (
            not isinstance(
                item,
                tuple,
            )
            or len(item) != 2
        ):
            raise ExcelReportValidationError(
                "Each row must be a two-item tuple."
            )

        label, value = item

        if not isinstance(
            label,
            str,
        ):
            raise TypeError(
                "Table labels must be strings."
            )

        normalized_label = label.strip()

        if not normalized_label:
            raise ExcelReportValidationError(
                "Table labels cannot be blank."
            )

        label_cell = worksheet.cell(
            row=current_row,
            column=1,
            value=normalized_label,
        )

        value_cell = worksheet.cell(
            row=current_row,
            column=2,
            value=_normalise_cell_value(
                value
            ),
        )

        _apply_label_style(
            label_cell
        )

        _apply_value_style(
            value_cell
        )

        current_row += 1

    return current_row


def _auto_fit_columns(
    worksheet: Worksheet,
    *,
    minimum_width: int = 12,
    maximum_width: int = 60,
    padding: int = 2,
) -> None:
    """
    Resize worksheet columns based on displayed content.
    """

    for parameter_name, value in (
        (
            "minimum_width",
            minimum_width,
        ),
        (
            "maximum_width",
            maximum_width,
        ),
        (
            "padding",
            padding,
        ),
    ):
        if isinstance(
            value,
            bool,
        ) or not isinstance(
            value,
            int,
        ):
            raise TypeError(
                f"{parameter_name} must be an integer."
            )

    if minimum_width <= 0:
        raise ExcelReportValidationError(
            "minimum_width must be greater than zero."
        )

    if maximum_width < minimum_width:
        raise ExcelReportValidationError(
            "maximum_width must be greater than or equal "
            "to minimum_width."
        )

    if padding < 0:
        raise ExcelReportValidationError(
            "padding cannot be negative."
        )

    for column_cells in worksheet.iter_cols():
        maximum_length = 0
        column_index: int | None = None

        for cell in column_cells:
            if column_index is None:
                column_index = cell.column

            if cell.value is None:
                continue

            display_value = str(
                cell.value
            )

            lines = (
                display_value.splitlines()
                or [display_value]
            )

            for line in lines:
                maximum_length = max(
                    maximum_length,
                    len(line),
                )

        if column_index is None:
            continue

        resolved_width = min(
            maximum_width,
            max(
                minimum_width,
                maximum_length + padding,
            ),
        )

        column_letter = get_column_letter(
            column_index
        )

        worksheet.column_dimensions[
            column_letter
        ].width = resolved_width


def _finalize_worksheet(
    worksheet: Worksheet,
    *,
    freeze_panes: str | None = None,
    auto_filter_range: str | None = None,
) -> None:
    """
    Apply standard finishing options to a worksheet.
    """

    worksheet.sheet_view.showGridLines = False

    if freeze_panes is not None:
        if not isinstance(
            freeze_panes,
            str,
        ):
            raise TypeError(
                "freeze_panes must be a string or None."
            )

        normalized_freeze_panes = (
            freeze_panes.strip()
        )

        if not normalized_freeze_panes:
            raise ExcelReportValidationError(
                "freeze_panes cannot be blank."
            )

        worksheet.freeze_panes = (
            normalized_freeze_panes
        )

    if auto_filter_range is not None:
        if not isinstance(
            auto_filter_range,
            str,
        ):
            raise TypeError(
                "auto_filter_range must be a string or None."
            )

        normalized_filter_range = (
            auto_filter_range.strip()
        )

        if not normalized_filter_range:
            raise ExcelReportValidationError(
                "auto_filter_range cannot be blank."
            )

        worksheet.auto_filter.ref = (
            normalized_filter_range
        )

    _auto_fit_columns(
        worksheet
    )


# ============================================================
# Worksheet Builders
# ============================================================


def _build_metadata_sheet(
    workbook: Workbook,
    report: PortfolioReport,
) -> None:
    """
    Build the report metadata worksheet.
    """

    worksheet = workbook.active
    worksheet.title = "Report Summary"

    _apply_title_style(
        worksheet,
        row=1,
        title=report.metadata.title,
        end_column=2,
    )

    _apply_section_style(
        worksheet,
        row=3,
        title="Report Information",
        end_column=2,
    )

    next_row = _write_key_value_table(
        worksheet,
        start_row=4,
        rows=[
            (
                "Application",
                report.metadata.application_name,
            ),
            (
                "Report Version",
                report.metadata.version,
            ),
            (
                "Generated At",
                report.metadata.generated_at,
            ),
            (
                "Report Title",
                report.metadata.title,
            ),
        ],
    )

    _apply_datetime_style(
        worksheet.cell(
            row=6,
            column=2,
        )
    )

    _apply_section_style(
        worksheet,
        row=next_row + 1,
        title="Included Report Sections",
        end_column=2,
    )

    _write_key_value_table(
        worksheet,
        start_row=next_row + 2,
        rows=[
            (
                "Portfolio Performance",
                "Included",
            ),
            (
                "Historical Analytics",
                (
                    "Included"
                    if report.history is not None
                    else UNAVAILABLE_VALUE
                ),
            ),
            (
                "Advanced Analytics",
                (
                    "Included"
                    if report.advanced_analytics is not None
                    else UNAVAILABLE_VALUE
                ),
            ),
            (
                "AI Portfolio Insights",
                (
                    "Included"
                    if report.ai_summary
                    else UNAVAILABLE_VALUE
                ),
            ),
            (
                "Notes",
                len(report.notes),
            ),
            (
                "Warnings",
                len(report.warnings),
            ),
        ],
    )

    _finalize_worksheet(
        worksheet,
        freeze_panes="A4",
    )


def _build_performance_sheet(
    workbook: Workbook,
    report: PortfolioReport,
) -> None:
    """
    Build the current portfolio performance worksheet.
    """

    worksheet = workbook.create_sheet(
        title="Portfolio Performance"
    )

    performance = report.performance

    _apply_title_style(
        worksheet,
        row=1,
        title="Portfolio Performance",
        end_column=2,
    )

    _apply_section_style(
        worksheet,
        row=3,
        title="Current Portfolio Metrics",
        end_column=2,
    )

    _write_key_value_table(
        worksheet,
        start_row=4,
        rows=[
            (
                "Total Investment",
                performance.total_investment,
            ),
            (
                "Current Portfolio Value",
                performance.current_value,
            ),
            (
                "Total Gain / Loss",
                performance.total_gain,
            ),
            (
                "Absolute Return",
                (
                    performance.absolute_return_percentage
                    / 100
                ),
            ),
            (
                "Total Holdings",
                performance.total_holdings,
            ),
            (
                "Profitable Holdings",
                performance.profitable_holdings,
            ),
            (
                "Loss-Making Holdings",
                performance.loss_making_holdings,
            ),
        ],
    )

    for row_number in (
        4,
        5,
        6,
    ):
        _apply_currency_style(
            worksheet.cell(
                row=row_number,
                column=2,
            )
        )

    _apply_percentage_style(
        worksheet.cell(
            row=7,
            column=2,
        )
    )

    for row_number in (
        8,
        9,
        10,
    ):
        worksheet.cell(
            row=row_number,
            column=2,
        ).alignment = Alignment(
            horizontal="right",
            vertical="top",
        )

    _finalize_worksheet(
        worksheet,
        freeze_panes="A4",
    )


def _build_history_sheet(
    workbook: Workbook,
    report: PortfolioReport,
) -> None:
    """
    Build the historical analytics worksheet.
    """

    worksheet = workbook.create_sheet(
        title="Historical Analytics"
    )

    _apply_title_style(
        worksheet,
        row=1,
        title="Historical Analytics",
        end_column=2,
    )

    history = report.history

    if history is None:
        _apply_section_style(
            worksheet,
            row=3,
            title="Availability",
            end_column=2,
        )

        _write_key_value_table(
            worksheet,
            start_row=4,
            rows=[
                (
                    "Status",
                    UNAVAILABLE_VALUE,
                ),
                (
                    "Message",
                    (
                        "Historical analytics were not available "
                        "when this report was generated."
                    ),
                ),
            ],
        )

        _finalize_worksheet(
            worksheet,
            freeze_panes="A4",
        )

        return

    _apply_section_style(
        worksheet,
        row=3,
        title="History Overview",
        end_column=2,
    )

    cagr_value = (
        history.cagr.cagr_percent / 100
        if history.cagr is not None
        else UNAVAILABLE_VALUE
    )

    drawdown_value = (
        history.drawdown.maximum_drawdown_percent
        / 100
        if history.drawdown is not None
        else UNAVAILABLE_VALUE
    )

    volatility_value = (
        history.volatility.annualised_volatility_percent
        / 100
        if history.volatility is not None
        else UNAVAILABLE_VALUE
    )

    _write_key_value_table(
        worksheet,
        start_row=4,
        rows=[
            (
                "First Snapshot",
                history.start_date,
            ),
            (
                "Latest Snapshot",
                history.end_date,
            ),
            (
                "Observations",
                history.observation_count,
            ),
            (
                "History Duration (Days)",
                history.duration_days,
            ),
            (
                "Starting Value",
                history.starting_value,
            ),
            (
                "Latest Value",
                history.latest_value,
            ),
            (
                "Lowest Value",
                history.minimum_value,
            ),
            (
                "Highest Value",
                history.maximum_value,
            ),
            (
                "Average Value",
                history.average_value,
            ),
            (
                "Absolute Growth",
                history.absolute_growth,
            ),
            (
                "Total Growth",
                history.total_growth_percent / 100,
            ),
            (
                "Historical CAGR",
                cagr_value,
            ),
            (
                "Maximum Drawdown",
                drawdown_value,
            ),
            (
                "Annualised Volatility",
                volatility_value,
            ),
        ],
    )

    for row_number in (
        4,
        5,
    ):
        _apply_date_style(
            worksheet.cell(
                row=row_number,
                column=2,
            )
        )

    for row_number in range(
        8,
        14,
    ):
        _apply_currency_style(
            worksheet.cell(
                row=row_number,
                column=2,
            )
        )

    for row_number in (
        14,
        15,
        16,
        17,
    ):
        cell = worksheet.cell(
            row=row_number,
            column=2,
        )

        if isinstance(
            cell.value,
            (int, float),
        ) and not isinstance(
            cell.value,
            bool,
        ):
            _apply_percentage_style(
                cell
            )

    _finalize_worksheet(
        worksheet,
        freeze_panes="A4",
    )


def _build_advanced_sheet(
    workbook: Workbook,
    report: PortfolioReport,
) -> None:
    """
    Build the advanced analytics worksheet.
    """

    worksheet = workbook.create_sheet(
        title="Advanced Analytics"
    )

    _apply_title_style(
        worksheet,
        row=1,
        title="Advanced Analytics",
        end_column=2,
    )

    advanced = report.advanced_analytics

    if advanced is None:
        _apply_section_style(
            worksheet,
            row=3,
            title="Availability",
            end_column=2,
        )

        _write_key_value_table(
            worksheet,
            start_row=4,
            rows=[
                (
                    "Status",
                    UNAVAILABLE_VALUE,
                ),
                (
                    "Message",
                    (
                        "Advanced analytics were not available "
                        "when this report was generated."
                    ),
                ),
            ],
        )

        _finalize_worksheet(
            worksheet,
            freeze_panes="A4",
        )

        return

    status = getattr(
        advanced,
        "status",
        UNAVAILABLE_VALUE,
    )

    available_metrics = getattr(
        advanced,
        "available_metrics",
        (),
    )

    unavailable_metrics = getattr(
        advanced,
        "unavailable_metrics",
        (),
    )

    failures = getattr(
        advanced,
        "failures",
        (),
    )

    _apply_section_style(
        worksheet,
        row=3,
        title="Analytics Availability",
        end_column=2,
    )

    next_row = _write_key_value_table(
        worksheet,
        start_row=4,
        rows=[
            (
                "Status",
                str(status).title(),
            ),
            (
                "Available Metrics",
                (
                    ", ".join(
                        str(metric)
                        for metric in available_metrics
                    )
                    if available_metrics
                    else UNAVAILABLE_VALUE
                ),
            ),
            (
                "Unavailable Metrics",
                (
                    ", ".join(
                        str(metric)
                        for metric in unavailable_metrics
                    )
                    if unavailable_metrics
                    else "None"
                ),
            ),
            (
                "Service Failures",
                (
                    len(failures)
                    if hasattr(
                        failures,
                        "__len__",
                    )
                    else UNAVAILABLE_VALUE
                ),
            ),
        ],
    )

    if failures:
        _apply_section_style(
            worksheet,
            row=next_row + 1,
            title="Failure Details",
            end_column=2,
        )

        failure_row = next_row + 2

        for index, failure in enumerate(
            failures,
            start=1,
        ):
            label_cell = worksheet.cell(
                row=failure_row,
                column=1,
                value=f"Failure {index}",
            )

            value_cell = worksheet.cell(
                row=failure_row,
                column=2,
                value=_safe_cell_text(
                    failure
                ),
            )

            _apply_label_style(
                label_cell
            )

            _apply_value_style(
                value_cell
            )

            value_cell.fill = WARNING_FILL
            value_cell.font = WARNING_FONT

            failure_row += 1

    _finalize_worksheet(
        worksheet,
        freeze_panes="A4",
    )


def _build_ai_sheet(
    workbook: Workbook,
    report: PortfolioReport,
) -> None:
    """
    Build the AI portfolio insights worksheet.
    """

    worksheet = workbook.create_sheet(
        title="AI Insights"
    )

    _apply_title_style(
        worksheet,
        row=1,
        title="AI Portfolio Insights",
        end_column=2,
    )

    _apply_section_style(
        worksheet,
        row=3,
        title="Generated Insights",
        end_column=2,
    )

    if not report.ai_summary:
        _write_key_value_table(
            worksheet,
            start_row=4,
            rows=[
                (
                    "Status",
                    UNAVAILABLE_VALUE,
                ),
                (
                    "Message",
                    (
                        "AI portfolio insights were not included "
                        "in this report."
                    ),
                ),
            ],
        )

        _finalize_worksheet(
            worksheet,
            freeze_panes="A4",
        )

        return

    rows = [
        (
            _safe_cell_text(
                key,
                maximum_length=200,
            ),
            _safe_cell_text(
                value
            ),
        )
        for key, value in report.ai_summary.items()
    ]

    _write_key_value_table(
        worksheet,
        start_row=4,
        rows=rows,
    )

    _finalize_worksheet(
        worksheet,
        freeze_panes="A4",
    )


def _build_notes_sheet(
    workbook: Workbook,
    report: PortfolioReport,
) -> None:
    """
    Build the notes, warnings, and disclaimer worksheet.
    """

    worksheet = workbook.create_sheet(
        title="Notes and Warnings"
    )

    _apply_title_style(
        worksheet,
        row=1,
        title="Notes and Warnings",
        end_column=2,
    )

    current_row = 3

    _apply_section_style(
        worksheet,
        row=current_row,
        title="Notes",
        end_column=2,
    )

    current_row += 1

    if report.notes:
        for index, note in enumerate(
            report.notes,
            start=1,
        ):
            label_cell = worksheet.cell(
                row=current_row,
                column=1,
                value=f"Note {index}",
            )

            value_cell = worksheet.cell(
                row=current_row,
                column=2,
                value=_safe_cell_text(
                    note
                ),
            )

            _apply_label_style(
                label_cell
            )

            _apply_value_style(
                value_cell
            )

            current_row += 1

    else:
        current_row = _write_key_value_table(
            worksheet,
            start_row=current_row,
            rows=[
                (
                    "Status",
                    "No additional notes were included.",
                )
            ],
        )

    current_row += 1

    _apply_section_style(
        worksheet,
        row=current_row,
        title="Warnings and Limitations",
        end_column=2,
    )

    current_row += 1

    if report.warnings:
        for index, warning in enumerate(
            report.warnings,
            start=1,
        ):
            label_cell = worksheet.cell(
                row=current_row,
                column=1,
                value=f"Warning {index}",
            )

            value_cell = worksheet.cell(
                row=current_row,
                column=2,
                value=_safe_cell_text(
                    warning
                ),
            )

            _apply_label_style(
                label_cell
            )

            value_cell.fill = WARNING_FILL
            value_cell.font = WARNING_FONT
            value_cell.border = THIN_BORDER
            value_cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

            current_row += 1

    else:
        current_row = _write_key_value_table(
            worksheet,
            start_row=current_row,
            rows=[
                (
                    "Status",
                    "No report warnings were generated.",
                )
            ],
        )

    current_row += 1

    _apply_section_style(
        worksheet,
        row=current_row,
        title="Important Disclaimer",
        end_column=2,
    )

    current_row += 1

    worksheet.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row + 2,
        end_column=2,
    )

    disclaimer_cell = worksheet.cell(
        row=current_row,
        column=1,
        value=(
            "This report is generated for informational and analytical "
            "purposes only. It does not constitute investment, tax, legal, "
            "or financial advice. Mutual fund investments are subject to "
            "market risks. Review scheme documents and consult a qualified "
            "financial professional before making investment decisions."
        ),
    )

    disclaimer_cell.font = SMALL_FONT
    disclaimer_cell.border = THIN_BORDER
    disclaimer_cell.alignment = Alignment(
        horizontal="left",
        vertical="top",
        wrap_text=True,
    )

    worksheet.row_dimensions[
        current_row
    ].height = 48

    _finalize_worksheet(
        worksheet,
        freeze_panes="A4",
    )


def _build_workbook(
    report: PortfolioReport,
) -> Workbook:
    """
    Build the complete portfolio-report workbook.
    """

    validated_report = _validate_report(
        report
    )

    workbook = _create_workbook(
        validated_report
    )

    _build_metadata_sheet(
        workbook,
        validated_report,
    )

    _build_performance_sheet(
        workbook,
        validated_report,
    )

    _build_history_sheet(
        workbook,
        validated_report,
    )

    _build_advanced_sheet(
        workbook,
        validated_report,
    )

    _build_ai_sheet(
        workbook,
        validated_report,
    )

    _build_notes_sheet(
        workbook,
        validated_report,
    )

    return workbook


# ============================================================
# Excel Report Service
# ============================================================


class ExcelReportService:
    """
    Generate professional portfolio reports as Excel workbooks.

    The service consumes a fully assembled PortfolioReport and provides
    workbook bytes, file-saving support, and download metadata.

    It performs no portfolio calculations and does not retrieve portfolio
    or analytics data.
    """

    @property
    def mime_type(self) -> str:
        """
        Return the Excel workbook MIME type.
        """

        return EXCEL_MIME_TYPE

    @property
    def default_filename(self) -> str:
        """
        Return the default downloadable Excel filename.
        """

        return DEFAULT_EXCEL_FILENAME

    def generate_bytes(
        self,
        report: PortfolioReport,
    ) -> bytes:
        """
        Generate a portfolio report as XLSX bytes.

        Args:
            report:
                Fully assembled PortfolioReport.

        Returns:
            Generated Excel workbook bytes.

        Raises:
            TypeError:
                When report is not a PortfolioReport.

            ExcelReportGenerationError:
                When workbook construction or serialization fails.
        """

        validated_report = _validate_report(
            report
        )

        workbook: Workbook | None = None
        buffer = BytesIO()
        excel_bytes = b""

        try:
            workbook = _build_workbook(
                validated_report
            )

            workbook.save(
                buffer
            )

            excel_bytes = buffer.getvalue()

        except ExcelReportError:
            raise

        except Exception as error:
            raise ExcelReportGenerationError(
                "Unable to generate the portfolio Excel report: "
                f"{error}"
            ) from error

        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass

            buffer.close()

        if not excel_bytes:
            raise ExcelReportGenerationError(
                "Generated Excel workbook is empty."
            )

        if not excel_bytes.startswith(
            b"PK"
        ):
            raise ExcelReportGenerationError(
                "Generated output is not a valid XLSX workbook."
            )

        return excel_bytes

    def save(
        self,
        report: PortfolioReport,
        output_path: str | Path,
        *,
        create_parent_directories: bool = True,
        overwrite: bool = False,
    ) -> Path:
        """
        Generate and save a portfolio Excel report.

        Args:
            report:
                Fully assembled PortfolioReport.

            output_path:
                Destination XLSX path.

            create_parent_directories:
                Whether missing parent directories should be created.

            overwrite:
                Whether an existing workbook may be replaced.

        Returns:
            Resolved output path.

        Raises:
            TypeError:
                When report, output_path, or boolean options are invalid.

            ExcelReportValidationError:
                When the output path is invalid, a parent directory is
                missing, or the file already exists and overwrite is False.

            ExcelReportGenerationError:
                When directory creation, workbook generation, or file writing
                fails.
        """

        validated_create_directories = _validate_boolean(
            create_parent_directories,
            parameter_name="create_parent_directories",
        )

        validated_overwrite = _validate_boolean(
            overwrite,
            parameter_name="overwrite",
        )

        path = _validate_output_path(
            output_path
        )

        if path.exists():
            if path.is_dir():
                raise ExcelReportValidationError(
                    "output_path must reference a file, not a directory."
                )

            if not validated_overwrite:
                raise ExcelReportValidationError(
                    "The output Excel workbook already exists. "
                    "Set overwrite=True to replace it."
                )

        parent = path.parent

        if parent.exists() and not parent.is_dir():
            raise ExcelReportValidationError(
                "The output parent path is not a directory."
            )

        if not parent.exists():
            if not validated_create_directories:
                raise ExcelReportValidationError(
                    "The output directory does not exist."
                )

            try:
                parent.mkdir(
                    parents=True,
                    exist_ok=True,
                )

            except Exception as error:
                raise ExcelReportGenerationError(
                    "Unable to create the output directory: "
                    f"{error}"
                ) from error

        excel_bytes = self.generate_bytes(
            report
        )

        try:
            path.write_bytes(
                excel_bytes
            )

        except Exception as error:
            raise ExcelReportGenerationError(
                "Unable to save the portfolio Excel report: "
                f"{error}"
            ) from error

        return path

    def prepare_download(
        self,
        report: PortfolioReport,
        *,
        filename: str = DEFAULT_EXCEL_FILENAME,
    ) -> tuple[bytes, str, str]:
        """
        Prepare workbook bytes and metadata for a Streamlit download button.

        Args:
            report:
                Fully assembled PortfolioReport.

            filename:
                Download filename.

        Returns:
            Tuple containing:

            - Excel workbook bytes
            - Validated filename
            - Excel workbook MIME type
        """

        resolved_filename = _validate_filename(
            filename
        )

        return (
            self.generate_bytes(
                report
            ),
            resolved_filename,
            self.mime_type,
        )


# ============================================================
# Convenience APIs
# ============================================================


def generate_portfolio_excel(
    report: PortfolioReport,
) -> bytes:
    """
    Generate a PortfolioReport as Excel workbook bytes.
    """

    service = ExcelReportService()

    return service.generate_bytes(
        report
    )


def save_portfolio_excel(
    report: PortfolioReport,
    output_path: str | Path,
    *,
    create_parent_directories: bool = True,
    overwrite: bool = False,
) -> Path:
    """
    Generate and save a PortfolioReport Excel workbook.
    """

    service = ExcelReportService()

    return service.save(
        report,
        output_path,
        create_parent_directories=create_parent_directories,
        overwrite=overwrite,
    )


# ============================================================
# Public Exports
# ============================================================

__all__ = [
    "DEFAULT_EXCEL_FILENAME",
    "EXCEL_MIME_TYPE",
    "MAX_CELL_TEXT_LENGTH",
    "ExcelReportError",
    "ExcelReportGenerationError",
    "ExcelReportService",
    "ExcelReportValidationError",
    "UNAVAILABLE_VALUE",
    "generate_portfolio_excel",
    "save_portfolio_excel",
]