"""Excel export service for FIFO tax-lot analytics."""

from __future__ import annotations

from io import BytesIO
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.worksheet.worksheet import Worksheet

from services.tax_lot_service import TaxLotAnalysis

DEFAULT_TAX_LOT_EXCEL_FILENAME: Final[str] = (
    "portfolio_tax_lots.xlsx"
)

TAX_LOT_EXCEL_MIME_TYPE: Final[str] = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)

_DATE_FORMAT: Final[str] = "dd mmm yyyy"
_UNITS_FORMAT: Final[str] = "0.000000"
_CURRENCY_FORMAT: Final[str] = (
    '₹#,##0.00;[Red]-₹#,##0.00'
)

_TITLE_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    fgColor="17365D",
)

_HEADER_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    fgColor="4472C4",
)

_OPEN_LOT_HEADERS: Final[tuple[str, ...]] = (
    "Portfolio ID",
    "Fund ID",
    "Source Transaction ID",
    "Acquisition Date",
    "Original Units",
    "Remaining Units",
    "Cost Per Unit",
    "Remaining Cost Basis",
)

_REALIZED_GAIN_HEADERS: Final[tuple[str, ...]] = (
    "Portfolio ID",
    "Fund ID",
    "Sell Transaction ID",
    "Source Buy Transaction ID",
    "Acquisition Date",
    "Disposal Date",
    "Holding Period Days",
    "Units Sold",
    "Sale Proceeds",
    "Cost Basis",
    "Realized Gain",
)


def _style_title(
    worksheet: Worksheet,
    *,
    title: str,
    end_column: int,
) -> None:
    """Apply a worksheet title."""

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=end_column,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value=title,
    )

    title_cell.fill = _TITLE_FILL
    title_cell.font = Font(
        name="Calibri",
        size=16,
        bold=True,
        color="FFFFFF",
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 28


def _write_headers(
    worksheet: Worksheet,
    *,
    headers: tuple[str, ...],
) -> None:
    """Write and style worksheet table headers."""

    for column, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=3,
            column=column,
            value=header,
        )

        cell.fill = _HEADER_FILL
        cell.font = Font(
            name="Calibri",
            size=10,
            bold=True,
            color="FFFFFF",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _configure_table(
    worksheet: Worksheet,
    *,
    end_column: str,
    data_row_count: int,
) -> None:
    """Configure shared table behavior."""

    final_row = max(
        3,
        data_row_count + 3,
    )

    worksheet.freeze_panes = "A4"
    worksheet.auto_filter.ref = (
        f"A3:{end_column}{final_row}"
    )
    worksheet.sheet_view.showGridLines = False


def _build_open_lots_sheet(
    workbook: Workbook,
    analysis: TaxLotAnalysis,
) -> None:
    """Build the open tax-lots worksheet."""

    worksheet = workbook.active
    worksheet.title = "Open Tax Lots"

    _style_title(
        worksheet,
        title="Open Tax Lots",
        end_column=len(_OPEN_LOT_HEADERS),
    )
    _write_headers(
        worksheet,
        headers=_OPEN_LOT_HEADERS,
    )

    for row_number, lot in enumerate(
        analysis.open_lots,
        start=4,
    ):
        values = (
            lot.portfolio_id,
            lot.fund_id,
            lot.source_transaction_id,
            lot.acquisition_date,
            float(lot.original_units),
            float(lot.remaining_units),
            float(lot.cost_per_unit),
            float(lot.remaining_cost_basis),
        )

        for column, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column,
                value=value,
            )
            cell.alignment = Alignment(
                vertical="top",
            )

        worksheet.cell(
            row=row_number,
            column=4,
        ).number_format = _DATE_FORMAT

        for column in (5, 6):
            worksheet.cell(
                row=row_number,
                column=column,
            ).number_format = _UNITS_FORMAT

        for column in (7, 8):
            worksheet.cell(
                row=row_number,
                column=column,
            ).number_format = _CURRENCY_FORMAT

    _configure_table(
        worksheet,
        end_column="H",
        data_row_count=len(analysis.open_lots),
    )

    column_widths = {
        "A": 14,
        "B": 12,
        "C": 24,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 16,
        "H": 24,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[
            column
        ].width = width


def _build_realized_gains_sheet(
    workbook: Workbook,
    analysis: TaxLotAnalysis,
) -> None:
    """Build the realized FIFO gains worksheet."""

    worksheet = workbook.create_sheet(
        title="Realized Gains"
    )

    _style_title(
        worksheet,
        title="Realized Gains",
        end_column=len(_REALIZED_GAIN_HEADERS),
    )
    _write_headers(
        worksheet,
        headers=_REALIZED_GAIN_HEADERS,
    )

    for row_number, match in enumerate(
        analysis.realized_gains,
        start=4,
    ):
        values = (
            match.portfolio_id,
            match.fund_id,
            match.sell_transaction_id,
            match.source_buy_transaction_id,
            match.acquisition_date,
            match.disposal_date,
            match.holding_period_days,
            float(match.units_sold),
            float(match.sale_proceeds),
            float(match.cost_basis),
            float(match.realized_gain),
        )

        for column, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column,
                value=value,
            )
            cell.alignment = Alignment(
                vertical="top",
            )

        for column in (5, 6):
            worksheet.cell(
                row=row_number,
                column=column,
            ).number_format = _DATE_FORMAT

        worksheet.cell(
            row=row_number,
            column=8,
        ).number_format = _UNITS_FORMAT

        for column in (9, 10, 11):
            worksheet.cell(
                row=row_number,
                column=column,
            ).number_format = _CURRENCY_FORMAT

    _configure_table(
        worksheet,
        end_column="K",
        data_row_count=len(analysis.realized_gains),
    )

    column_widths = {
        "A": 14,
        "B": 12,
        "C": 22,
        "D": 26,
        "E": 18,
        "F": 18,
        "G": 20,
        "H": 16,
        "I": 18,
        "J": 18,
        "K": 18,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[
            column
        ].width = width


class TaxLotExcelReportService:
    """Generate Excel reports from precomputed tax-lot analytics."""

    @property
    def mime_type(self) -> str:
        """Return the XLSX MIME type."""

        return TAX_LOT_EXCEL_MIME_TYPE

    @property
    def default_filename(self) -> str:
        """Return the default workbook filename."""

        return DEFAULT_TAX_LOT_EXCEL_FILENAME
    def prepare_download(
        self,
        analysis: TaxLotAnalysis,
    ) -> tuple[bytes, str, str]:
        """Return XLSX bytes, filename, and MIME type."""

        return (
            self.generate_bytes(analysis),
            self.default_filename,
            self.mime_type,
        )
    def generate_bytes(
        self,
        analysis: TaxLotAnalysis,
    ) -> bytes:
        """Return tax-lot analytics as XLSX bytes."""

        if not isinstance(
            analysis,
            TaxLotAnalysis,
        ):
            raise TypeError(
                "analysis must be a TaxLotAnalysis"
            )

        workbook = Workbook()
        buffer = BytesIO()

        try:
            workbook.properties.title = (
                "Portfolio Tax-Lot Analytics"
            )
            workbook.properties.creator = (
                "AI Mutual Fund Assistant"
            )

            _build_open_lots_sheet(
                workbook,
                analysis,
            )
            _build_realized_gains_sheet(
                workbook,
                analysis,
            )

            workbook.save(buffer)

            excel_bytes = buffer.getvalue()

        finally:
            workbook.close()
            buffer.close()

        return excel_bytes


__all__ = [
    "DEFAULT_TAX_LOT_EXCEL_FILENAME",
    "TAX_LOT_EXCEL_MIME_TYPE",
    "TaxLotExcelReportService",
]