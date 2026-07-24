"""Excel reporting for portfolio reconciliation results."""

from __future__ import annotations

from io import BytesIO
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.worksheet.worksheet import Worksheet

from services.portfolio_reconciliation_service import (
    PortfolioReconciliationResult,
)
DEFAULT_RECONCILIATION_EXCEL_FILENAME: Final[str] = (
    "portfolio_reconciliation.xlsx"
)

RECONCILIATION_EXCEL_MIME_TYPE: Final[str] = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)

_UNITS_FORMAT: Final[str] = "0.000000"
_CURRENCY_FORMAT: Final[str] = (
    '₹#,##0.00;[Red]-₹#,##0.00'
)

_TITLE_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    fgColor="17365D",
)

_HEADER_FILL: Final[PatternFill] = PatternFill(
    fill_type="solid",
    fgColor="4472C4",
)

_RECONCILIATION_HEADERS: Final[tuple[str, ...]] = (
    "Portfolio ID",
    "Fund ID",
    "Fund Name",
    "Position Units",
    "Transaction Units",
    "Unit Variance",
    "Position Cost Basis",
    "Transaction Cost Basis",
    "Cost Basis Variance",
    "Status",
)


def _style_title(
    worksheet: Worksheet,
) -> None:
    """Apply the reconciliation worksheet title."""

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(_RECONCILIATION_HEADERS),
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value="Portfolio Reconciliation",
    )

    title_cell.fill = _TITLE_FILL
    title_cell.font = Font(
        name="Calibri",
        size=16,
        bold=True,
        color="FFFFFF",
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 28


def _write_headers(
    worksheet: Worksheet,
) -> None:
    """Write and style reconciliation table headers."""

    for column, header in enumerate(
        _RECONCILIATION_HEADERS,
        start=1,
    ):
        cell = worksheet.cell(
            row=3,
            column=column,
            value=header,
        )

        cell.fill = _HEADER_FILL
        cell.font = Font(
            name="Calibri",
            size=10,
            bold=True,
            color="FFFFFF",
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

def _build_summary_sheet(
    workbook: Workbook,
    result: PortfolioReconciliationResult,
) -> None:
    """Build the aggregate reconciliation summary worksheet."""

    worksheet = workbook.create_sheet(
        title="Summary",
    )

    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=2,
    )

    title_cell = worksheet.cell(
        row=1,
        column=1,
        value="Reconciliation Summary",
    )

    title_cell.fill = _TITLE_FILL
    title_cell.font = Font(
        name="Calibri",
        size=16,
        bold=True,
        color="FFFFFF",
    )
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 28

    summary_rows = (
        (
            "Overall Status",
            (
                "Reconciled"
                if result.is_reconciled
                else "Unreconciled"
            ),
        ),
        ("Total Funds", result.total_count),
        ("Matched", result.matched_count),
        ("Unit Mismatches", result.unit_mismatch_count),
        (
            "Missing Positions",
            result.missing_position_count,
        ),
        (
            "Missing Tax Lots",
            result.missing_tax_lot_count,
        ),
        (
            "Cost-Basis Variances",
            result.cost_basis_variance_count,
        ),
    )

    for row_number, (label, value) in enumerate(
        summary_rows,
        start=3,
    ):
        label_cell = worksheet.cell(
            row=row_number,
            column=1,
            value=label,
        )
        value_cell = worksheet.cell(
            row=row_number,
            column=2,
            value=value,
        )

        label_cell.font = Font(
            name="Calibri",
            size=10,
            bold=True,
        )

        label_cell.alignment = Alignment(
            vertical="center",
        )
        value_cell.alignment = Alignment(
            vertical="center",
        )

    worksheet.column_dimensions["A"].width = 26
    worksheet.column_dimensions["B"].width = 18
    worksheet.sheet_view.showGridLines = False

def _build_reconciliation_sheet(
    workbook: Workbook,
    result: PortfolioReconciliationResult,
) -> None:
    """Build the detailed reconciliation worksheet."""

    worksheet = workbook.active
    worksheet.title = "Reconciliation"

    _style_title(worksheet)
    _write_headers(worksheet)

    for row_number, item in enumerate(
        result.items,
        start=4,
    ):
        values = (
            item.portfolio_id,
            item.fund_id,
            item.fund_name,
            float(item.position_units),
            float(item.transaction_units),
            float(item.unit_variance),
            float(item.position_cost_basis),
            float(item.transaction_cost_basis),
            float(item.cost_basis_variance),
            item.status,
        )

        for column, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column,
                value=value,
            )
            cell.alignment = Alignment(
                vertical="top",
            )

        for column in (4, 5, 6):
            worksheet.cell(
                row=row_number,
                column=column,
            ).number_format = _UNITS_FORMAT

        for column in (7, 8, 9):
            worksheet.cell(
                row=row_number,
                column=column,
            ).number_format = _CURRENCY_FORMAT

    final_row = max(
        3,
        len(result.items) + 3,
    )

    worksheet.freeze_panes = "A4"
    worksheet.auto_filter.ref = f"A3:J{final_row}"
    worksheet.sheet_view.showGridLines = False


class ReconciliationExcelReportService:
    """Generate Excel reports from reconciliation results."""
    @property
    def mime_type(self) -> str:
        """Return the XLSX MIME type."""

        return RECONCILIATION_EXCEL_MIME_TYPE

    @property
    def default_filename(self) -> str:
        """Return the default workbook filename."""

        return DEFAULT_RECONCILIATION_EXCEL_FILENAME

    def prepare_download(
        self,
        result: PortfolioReconciliationResult,
    ) -> tuple[bytes, str, str]:
        """Return XLSX bytes, filename, and MIME type."""

        return (
            self.generate_bytes(result),
            self.default_filename,
            self.mime_type,
        )
    def generate_bytes(
        self,
        result: PortfolioReconciliationResult,
    ) -> bytes:
        """Return reconciliation results as XLSX bytes."""

        if not isinstance(
            result,
            PortfolioReconciliationResult,
        ):
            raise TypeError(
                "result must be a PortfolioReconciliationResult"
            )

        workbook = Workbook()
        buffer = BytesIO()

        try:
            workbook.properties.title = (
                "Portfolio Reconciliation Report"
            )
            workbook.properties.creator = (
                "AI Mutual Fund Assistant"
            )

            _build_reconciliation_sheet(
                workbook,
                result,
            )
            _build_summary_sheet(
                workbook,
                result,
            )
            workbook.save(buffer)
            excel_bytes = buffer.getvalue()

        finally:
            workbook.close()
            buffer.close()

        return excel_bytes


__all__ = [
    "DEFAULT_RECONCILIATION_EXCEL_FILENAME",
    "RECONCILIATION_EXCEL_MIME_TYPE",
    "ReconciliationExcelReportService",
]