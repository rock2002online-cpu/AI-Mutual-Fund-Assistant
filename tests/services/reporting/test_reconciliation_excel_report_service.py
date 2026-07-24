"""Tests for portfolio reconciliation Excel reporting."""

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from services.portfolio_reconciliation_service import (
    PortfolioReconciliationItem,
    PortfolioReconciliationResult,
)
from services.reporting.reconciliation_excel_report_service import (
    ReconciliationExcelReportService,
)


def test_generate_bytes_builds_reconciliation_sheet() -> None:
    """Reconciliation items should be exported as a structured worksheet."""

    result = PortfolioReconciliationResult(
        items=[
            PortfolioReconciliationItem(
                portfolio_id=10,
                fund_id=20,
                fund_name="Example Equity Fund",
                position_units=Decimal("100.000000"),
                transaction_units=Decimal("100.000000"),
                unit_variance=Decimal("0.000000"),
                position_cost_basis=Decimal("1000.00"),
                transaction_cost_basis=Decimal("950.00"),
                cost_basis_variance=Decimal("50.00"),
                status="matched",
            )
        ],
        is_reconciled=True,
    )

    excel_bytes = (
        ReconciliationExcelReportService().generate_bytes(
            result
        )
    )

    workbook = load_workbook(
        BytesIO(excel_bytes),
    )

    try:
        assert "Reconciliation" in workbook.sheetnames

        worksheet = workbook["Reconciliation"]

        assert [
            worksheet.cell(row=3, column=column).value
            for column in range(1, 11)
        ] == [
            "Portfolio ID",
            "Fund ID",
            "Fund Name",
            "Position Units",
            "Transaction Units",
            "Unit Variance",
            "Position Cost Basis",
            "Transaction Cost Basis",
            "Cost Basis Variance",
            "Status",
        ]

        assert worksheet["A4"].value == 10
        assert worksheet["B4"].value == 20
        assert worksheet["C4"].value == "Example Equity Fund"
        assert worksheet["D4"].value == 100.0
        assert worksheet["E4"].value == 100.0
        assert worksheet["F4"].value == 0.0
        assert worksheet["G4"].value == 1000.0
        assert worksheet["H4"].value == 950.0
        assert worksheet["I4"].value == 50.0
        assert worksheet["J4"].value == "matched"

        assert worksheet.freeze_panes == "A4"
        assert worksheet.auto_filter.ref == "A3:J4"

    finally:
        workbook.close()
def test_generate_bytes_builds_reconciliation_summary_sheet() -> None:
    """The workbook should summarize authoritative unit reconciliation."""

    result = PortfolioReconciliationResult(
        items=[
            PortfolioReconciliationItem(
                portfolio_id=10,
                fund_id=20,
                fund_name="Matched Fund",
                position_units=Decimal("100.000000"),
                transaction_units=Decimal("100.000000"),
                unit_variance=Decimal("0.000000"),
                position_cost_basis=Decimal("1000.00"),
                transaction_cost_basis=Decimal("950.00"),
                cost_basis_variance=Decimal("50.00"),
                status="matched",
            ),
            PortfolioReconciliationItem(
                portfolio_id=10,
                fund_id=21,
                fund_name="Unit Mismatch Fund",
                position_units=Decimal("80.000000"),
                transaction_units=Decimal("75.000000"),
                unit_variance=Decimal("5.000000"),
                position_cost_basis=Decimal("800.00"),
                transaction_cost_basis=Decimal("750.00"),
                cost_basis_variance=Decimal("50.00"),
                status="unit_mismatch",
            ),
            PortfolioReconciliationItem(
                portfolio_id=10,
                fund_id=22,
                fund_name=None,
                position_units=Decimal("0.000000"),
                transaction_units=Decimal("25.000000"),
                unit_variance=Decimal("-25.000000"),
                position_cost_basis=Decimal("0.00"),
                transaction_cost_basis=Decimal("250.00"),
                cost_basis_variance=Decimal("-250.00"),
                status="missing_position",
            ),
            PortfolioReconciliationItem(
                portfolio_id=10,
                fund_id=23,
                fund_name="Missing Tax Lots Fund",
                position_units=Decimal("30.000000"),
                transaction_units=Decimal("0.000000"),
                unit_variance=Decimal("30.000000"),
                position_cost_basis=Decimal("300.00"),
                transaction_cost_basis=Decimal("0.00"),
                cost_basis_variance=Decimal("300.00"),
                status="missing_tax_lots",
            ),
        ],
        is_reconciled=False,
    )

    excel_bytes = (
        ReconciliationExcelReportService().generate_bytes(
            result
        )
    )

    workbook = load_workbook(
        BytesIO(excel_bytes),
    )

    try:
        assert "Summary" in workbook.sheetnames

        worksheet = workbook["Summary"]

        assert worksheet["A3"].value == "Overall Status"
        assert worksheet["B3"].value == "Unreconciled"
        assert worksheet["A4"].value == "Total Funds"
        assert worksheet["B4"].value == 4
        assert worksheet["A5"].value == "Matched"
        assert worksheet["B5"].value == 1
        assert worksheet["A6"].value == "Unit Mismatches"
        assert worksheet["B6"].value == 1
        assert worksheet["A7"].value == "Missing Positions"
        assert worksheet["B7"].value == 1
        assert worksheet["A8"].value == "Missing Tax Lots"
        assert worksheet["B8"].value == 1
        assert worksheet["A9"].value == "Cost-Basis Variances"
        assert worksheet["B9"].value == 1

    finally:
        workbook.close()
def test_prepare_download_returns_bytes_and_metadata() -> None:
    """Download preparation should return XLSX data and metadata."""

    result = PortfolioReconciliationResult(
        items=[],
        is_reconciled=True,
    )

    service = ReconciliationExcelReportService()

    excel_bytes, filename, mime_type = (
        service.prepare_download(
            result
        )
    )

    assert excel_bytes.startswith(b"PK")
    assert filename == "portfolio_reconciliation.xlsx"
    assert mime_type == (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )
def test_generate_bytes_sets_workbook_metadata() -> None:
    """The reconciliation workbook should identify its purpose and creator."""

    result = PortfolioReconciliationResult(
        items=[],
        is_reconciled=True,
    )

    excel_bytes = (
        ReconciliationExcelReportService().generate_bytes(
            result
        )
    )

    workbook = load_workbook(
        BytesIO(excel_bytes),
    )

    try:
        assert (
            workbook.properties.title
            == "Portfolio Reconciliation Report"
        )
        assert (
            workbook.properties.creator
            == "AI Mutual Fund Assistant"
        )

    finally:
        workbook.close()