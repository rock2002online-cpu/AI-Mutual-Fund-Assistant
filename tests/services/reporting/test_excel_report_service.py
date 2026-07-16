"""
Unit tests for the professional Excel portfolio reporting service.

Coverage includes:

- Report and filename validation
- Output-path validation
- Boolean validation
- Safe cell-text formatting
- Workbook construction
- Worksheet generation
- Excel byte generation
- Invalid XLSX detection
- Error wrapping
- File-saving behavior
- Download preparation
- Convenience APIs

The tests isolate Excel reporting from portfolio calculations and external
analytics services.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import Mock

import pytest
from openpyxl import Workbook, load_workbook

import services.reporting.excel_report_service as excel_report_module

from services.reporting.excel_report_service import (
    DEFAULT_EXCEL_FILENAME,
    EXCEL_MIME_TYPE,
    MAX_CELL_TEXT_LENGTH,
    UNAVAILABLE_VALUE,
    ExcelReportGenerationError,
    ExcelReportService,
    ExcelReportValidationError,
    _auto_fit_columns,
    _build_workbook,
    _create_workbook,
    _finalize_worksheet,
    _normalise_cell_value,
    _safe_cell_text,
    _validate_boolean,
    _validate_filename,
    _validate_output_path,
    _validate_report,
    generate_portfolio_excel,
    save_portfolio_excel,
)


# ============================================================
# Test Report Doubles
# ============================================================


class FakePortfolioReport:
    """
    Lightweight report object matching the Excel service contract.
    """

    def __init__(
        self,
        *,
        include_history: bool = True,
        include_advanced: bool = True,
        include_ai: bool = True,
        include_notes: bool = True,
        include_warnings: bool = True,
    ) -> None:
        self.metadata = SimpleNamespace(
            title="Portfolio Analytics Report",
            application_name="AI Mutual Fund Assistant",
            version="8.0.0",
            generated_at=datetime(
                2026,
                7,
                16,
                12,
                30,
            ),
        )

        self.performance = SimpleNamespace(
            total_investment=100_000.0,
            current_value=125_000.0,
            total_gain=25_000.0,
            absolute_return_percentage=25.0,
            total_holdings=5,
            profitable_holdings=4,
            loss_making_holdings=1,
        )

        self.history = (
            SimpleNamespace(
                start_date=date(
                    2024,
                    1,
                    1,
                ),
                end_date=date(
                    2026,
                    1,
                    1,
                ),
                observation_count=25,
                duration_days=731,
                starting_value=90_000.0,
                latest_value=125_000.0,
                minimum_value=85_000.0,
                maximum_value=130_000.0,
                average_value=108_000.0,
                absolute_growth=35_000.0,
                total_growth_percent=38.89,
                cagr=SimpleNamespace(
                    cagr_percent=17.85,
                ),
                drawdown=SimpleNamespace(
                    maximum_drawdown_percent=-12.5,
                ),
                volatility=SimpleNamespace(
                    annualised_volatility_percent=14.25,
                ),
            )
            if include_history
            else None
        )

        self.advanced_analytics = (
            SimpleNamespace(
                status="complete",
                available_metrics=(
                    "xirr",
                    "sharpe_ratio",
                    "alpha",
                ),
                unavailable_metrics=(),
                failures=(),
            )
            if include_advanced
            else None
        )

        self.ai_summary = (
            {
                "Executive Summary": (
                    "Portfolio performance remains strong."
                ),
                "Risk Level": "Moderate",
                "Recommended Action": "Review allocation quarterly.",
            }
            if include_ai
            else {}
        )

        self.notes = (
            (
                "Values are based on the latest available NAV.",
                "Historical analytics depend on available snapshots.",
            )
            if include_notes
            else ()
        )

        self.warnings = (
            (
                "Past performance does not guarantee future returns.",
            )
            if include_warnings
            else ()
        )


@pytest.fixture
def fake_report(
    monkeypatch: pytest.MonkeyPatch,
) -> FakePortfolioReport:
    """
    Return a complete fake report accepted by service validation.
    """

    monkeypatch.setattr(
        excel_report_module,
        "PortfolioReport",
        FakePortfolioReport,
    )

    return FakePortfolioReport()


@pytest.fixture
def minimal_fake_report(
    monkeypatch: pytest.MonkeyPatch,
) -> FakePortfolioReport:
    """
    Return a report without optional reporting sections.
    """

    monkeypatch.setattr(
        excel_report_module,
        "PortfolioReport",
        FakePortfolioReport,
    )

    return FakePortfolioReport(
        include_history=False,
        include_advanced=False,
        include_ai=False,
        include_notes=False,
        include_warnings=False,
    )


# ============================================================
# Report Validation
# ============================================================


def test_validate_report_accepts_portfolio_report(
    fake_report: FakePortfolioReport,
) -> None:
    assert (
        _validate_report(
            fake_report  # type: ignore[arg-type]
        )
        is fake_report
    )


@pytest.mark.parametrize(
    "invalid_report",
    [
        None,
        {},
        [],
        "report",
        123,
        object(),
    ],
)
def test_validate_report_rejects_invalid_type(
    invalid_report: object,
) -> None:
    with pytest.raises(
        TypeError,
        match="report must be a PortfolioReport",
    ):
        _validate_report(
            invalid_report  # type: ignore[arg-type]
        )


# ============================================================
# Output Path Validation
# ============================================================


@pytest.mark.parametrize(
    "value",
    [
        "portfolio.xlsx",
        "PORTFOLIO.XLSX",
        Path("reports.xlsx"),
    ],
)
def test_validate_output_path_accepts_xlsx_path(
    value: str | Path,
) -> None:
    result = _validate_output_path(value)

    assert isinstance(
        result,
        Path,
    )

    assert result.is_absolute()
    assert result.suffix.lower() == ".xlsx"


@pytest.mark.parametrize(
    "value",
    [
        None,
        1,
        1.5,
        object(),
    ],
)
def test_validate_output_path_rejects_invalid_type(
    value: object,
) -> None:
    with pytest.raises(
        TypeError,
        match="output_path must be a string or Path",
    ):
        _validate_output_path(
            value  # type: ignore[arg-type]
        )


@pytest.mark.parametrize(
    "value",
    [
        "",
        " ",
        "\t",
        "\n",
    ],
)
def test_validate_output_path_rejects_blank_path(
    value: str,
) -> None:
    with pytest.raises(
        ExcelReportValidationError,
        match="output_path cannot be blank",
    ):
        _validate_output_path(value)


@pytest.mark.parametrize(
    "value",
    [
        "portfolio",
        "portfolio.xls",
        "portfolio.csv",
        "portfolio.pdf",
    ],
)
def test_validate_output_path_rejects_non_xlsx_extension(
    value: str,
) -> None:
    with pytest.raises(
        ExcelReportValidationError,
        match="output_path must end with .xlsx",
    ):
        _validate_output_path(value)


# ============================================================
# Filename Validation
# ============================================================


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (
            "portfolio.xlsx",
            "portfolio.xlsx",
        ),
        (
            " PORTFOLIO.XLSX ",
            "PORTFOLIO.XLSX",
        ),
        (
            "portfolio report.xlsx",
            "portfolio report.xlsx",
        ),
    ],
)
def test_validate_filename_accepts_valid_filename(
    value: str,
    expected: str,
) -> None:
    assert _validate_filename(value) == expected


@pytest.mark.parametrize(
    "value",
    [
        None,
        123,
        Path("portfolio.xlsx"),
        object(),
    ],
)
def test_validate_filename_rejects_invalid_type(
    value: object,
) -> None:
    with pytest.raises(
        TypeError,
        match="filename must be a string",
    ):
        _validate_filename(
            value  # type: ignore[arg-type]
        )


@pytest.mark.parametrize(
    "value",
    [
        "",
        " ",
        "\t",
        "\n",
    ],
)
def test_validate_filename_rejects_blank_filename(
    value: str,
) -> None:
    with pytest.raises(
        ExcelReportValidationError,
        match="filename cannot be blank",
    ):
        _validate_filename(value)


@pytest.mark.parametrize(
    "value",
    [
        "reports/portfolio.xlsx",
        "reports\\portfolio.xlsx",
        "../portfolio.xlsx",
        "./portfolio.xlsx",
    ],
)
def test_validate_filename_rejects_directory_components(
    value: str,
) -> None:
    with pytest.raises(
        ExcelReportValidationError,
        match="filename must not contain directory components",
    ):
        _validate_filename(value)


@pytest.mark.parametrize(
    "value",
    [
        "portfolio",
        "portfolio.xls",
        "portfolio.csv",
        "portfolio.xlsx.txt",
    ],
)
def test_validate_filename_rejects_non_xlsx_filename(
    value: str,
) -> None:
    with pytest.raises(
        ExcelReportValidationError,
        match="filename must end with .xlsx",
    ):
        _validate_filename(value)


# ============================================================
# Boolean Validation
# ============================================================


@pytest.mark.parametrize(
    "value",
    [
        True,
        False,
    ],
)
def test_validate_boolean_accepts_boolean(
    value: bool,
) -> None:
    assert (
        _validate_boolean(
            value,
            parameter_name="overwrite",
        )
        is value
    )


@pytest.mark.parametrize(
    "value",
    [
        None,
        0,
        1,
        "true",
        object(),
    ],
)
def test_validate_boolean_rejects_non_boolean(
    value: object,
) -> None:
    with pytest.raises(
        TypeError,
        match="overwrite must be a boolean",
    ):
        _validate_boolean(
            value,  # type: ignore[arg-type]
            parameter_name="overwrite",
        )


@pytest.mark.parametrize(
    "parameter_name",
    [
        "",
        " ",
        "\t",
    ],
)
def test_validate_boolean_rejects_blank_parameter_name(
    parameter_name: str,
) -> None:
    with pytest.raises(
        ExcelReportValidationError,
        match="parameter_name cannot be blank",
    ):
        _validate_boolean(
            True,
            parameter_name=parameter_name,
        )


def test_validate_boolean_rejects_invalid_parameter_name_type() -> None:
    with pytest.raises(
        TypeError,
        match="parameter_name must be a string",
    ):
        _validate_boolean(
            True,
            parameter_name=None,  # type: ignore[arg-type]
        )


# ============================================================
# Safe Cell Text
# ============================================================


def test_safe_cell_text_handles_none() -> None:
    assert _safe_cell_text(None) == UNAVAILABLE_VALUE


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (
            True,
            "Yes",
        ),
        (
            False,
            "No",
        ),
    ],
)
def test_safe_cell_text_handles_boolean(
    value: bool,
    expected: str,
) -> None:
    assert _safe_cell_text(value) == expected


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (
            0.0,
            "0.0000",
        ),
        (
            12.5,
            "12.5000",
        ),
        (
            1_234.56789,
            "1,234.5679",
        ),
    ],
)
def test_safe_cell_text_formats_finite_float(
    value: float,
    expected: str,
) -> None:
    assert _safe_cell_text(value) == expected


@pytest.mark.parametrize(
    "value",
    [
        float("nan"),
        float("inf"),
        float("-inf"),
    ],
)
def test_safe_cell_text_handles_non_finite_float(
    value: float,
) -> None:
    assert _safe_cell_text(value) == UNAVAILABLE_VALUE


def test_safe_cell_text_handles_mapping() -> None:
    result = _safe_cell_text(
        {
            "status": "complete",
            "count": 3,
        }
    )

    assert result == (
        "status: complete; count: 3"
    )


def test_safe_cell_text_handles_sequence() -> None:
    result = _safe_cell_text(
        [
            "alpha",
            None,
            True,
        ]
    )

    assert result == (
        "alpha; Unavailable; Yes"
    )


def test_safe_cell_text_normalizes_whitespace() -> None:
    result = _safe_cell_text(
        " alpha\n beta\t gamma "
    )

    assert result == "alpha beta gamma"


def test_safe_cell_text_truncates_long_content() -> None:
    result = _safe_cell_text(
        "abcdefghij",
        maximum_length=8,
    )

    assert result == "abcde..."
    assert len(result) == 8


def test_safe_cell_text_supports_tiny_maximum_length() -> None:
    assert (
        _safe_cell_text(
            "abcdef",
            maximum_length=2,
        )
        == "ab"
    )


@pytest.mark.parametrize(
    "value",
    [
        True,
        False,
        None,
        1.5,
        "10",
    ],
)
def test_safe_cell_text_rejects_invalid_maximum_length_type(
    value: object,
) -> None:
    with pytest.raises(
        TypeError,
        match="maximum_length must be an integer",
    ):
        _safe_cell_text(
            "text",
            maximum_length=value,  # type: ignore[arg-type]
        )


@pytest.mark.parametrize(
    "value",
    [
        0,
        -1,
        -100,
    ],
)
def test_safe_cell_text_rejects_non_positive_maximum_length(
    value: int,
) -> None:
    with pytest.raises(
        ExcelReportValidationError,
        match="maximum_length must be greater than zero",
    ):
        _safe_cell_text(
            "text",
            maximum_length=value,
        )


def test_default_cell_text_limit_is_excel_safe() -> None:
    result = _safe_cell_text(
        "a" * (
            MAX_CELL_TEXT_LENGTH + 100
        )
    )

    assert len(result) == MAX_CELL_TEXT_LENGTH
    assert result.endswith("...")


# ============================================================
# Cell Value Normalisation
# ============================================================

def test_normalise_cell_value_preserves_naive_datetime() -> None:
    value = datetime(
        2026,
        7,
        16,
        14,
        30,
    )

    result = _normalise_cell_value(
        value
    )

    assert result == value
    assert result.tzinfo is None


def test_normalise_cell_value_converts_utc_datetime_to_naive() -> None:
    value = datetime(
        2026,
        7,
        16,
        14,
        30,
        tzinfo=timezone.utc,
    )

    result = _normalise_cell_value(
        value
    )

    assert result == datetime(
        2026,
        7,
        16,
        14,
        30,
    )

    assert result.tzinfo is None


def test_normalise_cell_value_converts_offset_datetime_to_naive_utc() -> None:
    indian_standard_time = timezone(
        timedelta(
            hours=5,
            minutes=30,
        )
    )

    value = datetime(
        2026,
        7,
        16,
        20,
        0,
        tzinfo=indian_standard_time,
    )

    result = _normalise_cell_value(
        value
    )

    assert result == datetime(
        2026,
        7,
        16,
        14,
        30,
    )

    assert result.tzinfo is None


def test_generate_bytes_supports_timezone_aware_generated_at(
    fake_report: FakePortfolioReport,
) -> None:
    fake_report.metadata.generated_at = datetime(
        2026,
        7,
        16,
        14,
        30,
        tzinfo=timezone.utc,
    )

    service = ExcelReportService()

    result = service.generate_bytes(
        fake_report  # type: ignore[arg-type]
    )

    assert isinstance(
        result,
        bytes,
    )

    assert result.startswith(
        b"PK"
    )
    
@pytest.mark.parametrize(
    "value",
    [
        True,
        False,
        0,
        10,
        1.5,
        "text",
        date(
            2026,
            1,
            1,
        ),
        datetime(
            2026,
            1,
            1,
            12,
            30,
        ),
    ],
)
def test_normalise_cell_value_preserves_supported_values(
    value: object,
) -> None:
    result = _normalise_cell_value(value)

    assert result == value


@pytest.mark.parametrize(
    "value",
    [
        float("nan"),
        float("inf"),
        float("-inf"),
        None,
    ],
)
def test_normalise_cell_value_replaces_unavailable_values(
    value: object,
) -> None:
    assert (
        _normalise_cell_value(value)
        == UNAVAILABLE_VALUE
    )


def test_normalise_cell_value_converts_complex_object() -> None:
    result = _normalise_cell_value(
        {
            "metric": "xirr",
            "status": "complete",
        }
    )

    assert result == (
        "metric: xirr; status: complete"
    )


# ============================================================
# Workbook Creation
# ============================================================


def test_create_workbook_sets_document_properties(
    fake_report: FakePortfolioReport,
) -> None:
    workbook = _create_workbook(
        fake_report  # type: ignore[arg-type]
    )

    try:
        assert (
            workbook.properties.title
            == fake_report.metadata.title
        )

        assert (
            workbook.properties.creator
            == fake_report.metadata.application_name
        )

        assert (
            workbook.properties.subject
            == (
                "Professional mutual fund "
                "portfolio analytics report"
            )
        )

    finally:
        workbook.close()


def test_build_workbook_creates_expected_worksheets(
    fake_report: FakePortfolioReport,
) -> None:
    workbook = _build_workbook(
        fake_report  # type: ignore[arg-type]
    )

    try:
        assert workbook.sheetnames == [
            "Report Summary",
            "Portfolio Performance",
            "Historical Analytics",
            "Advanced Analytics",
            "AI Insights",
            "Notes and Warnings",
        ]

    finally:
        workbook.close()


def test_build_workbook_populates_performance_values(
    fake_report: FakePortfolioReport,
) -> None:
    workbook = _build_workbook(
        fake_report  # type: ignore[arg-type]
    )

    try:
        worksheet = workbook[
            "Portfolio Performance"
        ]

        assert worksheet["A4"].value == (
            "Total Investment"
        )

        assert worksheet["B4"].value == (
            100_000.0
        )

        assert worksheet["A7"].value == (
            "Absolute Return"
        )

        assert worksheet["B7"].value == (
            0.25
        )

        assert "%" in worksheet[
            "B7"
        ].number_format

    finally:
        workbook.close()


def test_build_workbook_populates_history_sheet(
    fake_report: FakePortfolioReport,
) -> None:
    workbook = _build_workbook(
        fake_report  # type: ignore[arg-type]
    )

    try:
        worksheet = workbook[
            "Historical Analytics"
        ]

        assert worksheet["A4"].value == (
            "First Snapshot"
        )

        assert worksheet["B4"].value == date(
            2024,
            1,
            1,
        )

        assert worksheet["A14"].value == (
            "Total Growth"
        )

        assert worksheet["B14"].value == pytest.approx(
            0.3889
        )

    finally:
        workbook.close()


def test_build_workbook_handles_unavailable_optional_sections(
    minimal_fake_report: FakePortfolioReport,
) -> None:
    workbook = _build_workbook(
        minimal_fake_report  # type: ignore[arg-type]
    )

    try:
        history_sheet = workbook[
            "Historical Analytics"
        ]

        advanced_sheet = workbook[
            "Advanced Analytics"
        ]

        ai_sheet = workbook[
            "AI Insights"
        ]

        assert history_sheet["B4"].value == (
            UNAVAILABLE_VALUE
        )

        assert advanced_sheet["B4"].value == (
            UNAVAILABLE_VALUE
        )

        assert ai_sheet["B4"].value == (
            UNAVAILABLE_VALUE
        )

    finally:
        workbook.close()


def test_workbook_worksheets_hide_gridlines(
    fake_report: FakePortfolioReport,
) -> None:
    workbook = _build_workbook(
        fake_report  # type: ignore[arg-type]
    )

    try:
        for worksheet in workbook.worksheets:
            assert (
                worksheet.sheet_view.showGridLines
                is False
            )

    finally:
        workbook.close()


# ============================================================
# Worksheet Utilities
# ============================================================


def test_auto_fit_columns_sets_column_widths() -> None:
    workbook = Workbook()
    worksheet = workbook.active

    worksheet["A1"] = "Short"
    worksheet["B1"] = (
        "This is a substantially longer value"
    )

    _auto_fit_columns(
        worksheet
    )

    try:
        assert (
            worksheet.column_dimensions["A"].width
            >= 12
        )

        assert (
            worksheet.column_dimensions["B"].width
            > worksheet.column_dimensions["A"].width
        )

    finally:
        workbook.close()


@pytest.mark.parametrize(
    "kwargs",
    [
        {
            "minimum_width": 0,
        },
        {
            "minimum_width": 20,
            "maximum_width": 10,
        },
        {
            "padding": -1,
        },
    ],
)
def test_auto_fit_columns_rejects_invalid_ranges(
    kwargs: dict[str, int],
) -> None:
    workbook = Workbook()

    try:
        with pytest.raises(
            ExcelReportValidationError,
        ):
            _auto_fit_columns(
                workbook.active,
                **kwargs,
            )

    finally:
        workbook.close()


def test_finalize_worksheet_sets_freeze_panes_and_filter() -> None:
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append(
        [
            "Metric",
            "Value",
        ]
    )

    _finalize_worksheet(
        worksheet,
        freeze_panes="A2",
        auto_filter_range="A1:B1",
    )

    try:
        assert worksheet.freeze_panes == "A2"
        assert worksheet.auto_filter.ref == "A1:B1"
        assert (
            worksheet.sheet_view.showGridLines
            is False
        )

    finally:
        workbook.close()


# ============================================================
# Service Properties
# ============================================================


def test_excel_report_service_returns_mime_type() -> None:
    service = ExcelReportService()

    assert service.mime_type == EXCEL_MIME_TYPE

    assert service.mime_type == (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )


def test_excel_report_service_returns_default_filename() -> None:
    service = ExcelReportService()

    assert (
        service.default_filename
        == DEFAULT_EXCEL_FILENAME
    )

    assert (
        service.default_filename
        == "portfolio_report.xlsx"
    )


# ============================================================
# Workbook Byte Generation
# ============================================================


def test_generate_bytes_returns_valid_xlsx_workbook(
    fake_report: FakePortfolioReport,
) -> None:
    service = ExcelReportService()

    result = service.generate_bytes(
        fake_report  # type: ignore[arg-type]
    )

    assert isinstance(
        result,
        bytes,
    )

    assert result.startswith(
        b"PK"
    )

    assert len(result) > 1_000


def test_generate_bytes_can_be_loaded_by_openpyxl(
    fake_report: FakePortfolioReport,
) -> None:
    service = ExcelReportService()

    result = service.generate_bytes(
        fake_report  # type: ignore[arg-type]
    )

    workbook = load_workbook(
        BytesIO(result),
        data_only=False,
    )

    try:
        assert workbook.sheetnames == [
            "Report Summary",
            "Portfolio Performance",
            "Historical Analytics",
            "Advanced Analytics",
            "AI Insights",
            "Notes and Warnings",
        ]

        assert (
            workbook[
                "Portfolio Performance"
            ]["B4"].value
            == 100_000
        )

    finally:
        workbook.close()


@pytest.mark.parametrize(
    "invalid_report",
    [
        None,
        {},
        [],
        "report",
        123,
        object(),
    ],
)
def test_generate_bytes_rejects_invalid_report(
    invalid_report: object,
) -> None:
    service = ExcelReportService()

    with pytest.raises(
        TypeError,
        match="report must be a PortfolioReport",
    ):
        service.generate_bytes(
            invalid_report  # type: ignore[arg-type]
        )


def test_generate_bytes_wraps_workbook_build_failure(
    fake_report: FakePortfolioReport,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        excel_report_module,
        "_build_workbook",
        Mock(
            side_effect=RuntimeError(
                "simulated workbook failure"
            )
        ),
    )

    service = ExcelReportService()

    with pytest.raises(
        ExcelReportGenerationError,
        match=(
            "Unable to generate the portfolio Excel report: "
            "simulated workbook failure"
        ),
    ) as error_info:
        service.generate_bytes(
            fake_report  # type: ignore[arg-type]
        )

    assert isinstance(
        error_info.value.__cause__,
        RuntimeError,
    )


def test_generate_bytes_rejects_empty_workbook_output(
    fake_report: FakePortfolioReport,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class EmptyWorkbook:
        def save(
            self,
            buffer: BytesIO,
        ) -> None:
            return None

        def close(self) -> None:
            return None

    monkeypatch.setattr(
        excel_report_module,
        "_build_workbook",
        lambda report: EmptyWorkbook(),
    )

    service = ExcelReportService()

    with pytest.raises(
        ExcelReportGenerationError,
        match="Generated Excel workbook is empty",
    ):
        service.generate_bytes(
            fake_report  # type: ignore[arg-type]
        )


def test_generate_bytes_rejects_invalid_xlsx_signature(
    fake_report: FakePortfolioReport,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    class InvalidWorkbook:
        def save(
            self,
            buffer: BytesIO,
        ) -> None:
            buffer.write(
                b"invalid-workbook"
            )

        def close(self) -> None:
            return None

    monkeypatch.setattr(
        excel_report_module,
        "_build_workbook",
        lambda report: InvalidWorkbook(),
    )

    service = ExcelReportService()

    with pytest.raises(
        ExcelReportGenerationError,
        match=(
            "Generated output is not a valid XLSX workbook"
        ),
    ):
        service.generate_bytes(
            fake_report  # type: ignore[arg-type]
        )


# ============================================================
# File Saving
# ============================================================


def test_save_writes_excel_file(
    fake_report: FakePortfolioReport,
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "portfolio-report.xlsx"
    )

    service = ExcelReportService()

    result = service.save(
        fake_report,  # type: ignore[arg-type]
        output_path,
    )

    assert result == output_path.resolve()
    assert result.exists()
    assert result.read_bytes().startswith(
        b"PK"
    )


def test_save_creates_parent_directories(
    fake_report: FakePortfolioReport,
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "nested"
        / "reports"
        / "portfolio.xlsx"
    )

    service = ExcelReportService()

    result = service.save(
        fake_report,  # type: ignore[arg-type]
        output_path,
    )

    assert result.exists()
    assert result.parent.exists()


def test_save_rejects_missing_parent_when_creation_disabled(
    fake_report: FakePortfolioReport,
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "missing"
        / "portfolio.xlsx"
    )

    service = ExcelReportService()

    with pytest.raises(
        ExcelReportValidationError,
        match="output directory does not exist",
    ):
        service.save(
            fake_report,  # type: ignore[arg-type]
            output_path,
            create_parent_directories=False,
        )


def test_save_rejects_existing_file_without_overwrite(
    fake_report: FakePortfolioReport,
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "portfolio.xlsx"
    )

    output_path.write_bytes(
        b"existing"
    )

    service = ExcelReportService()

    with pytest.raises(
        ExcelReportValidationError,
        match=(
            "output Excel workbook already exists"
        ),
    ):
        service.save(
            fake_report,  # type: ignore[arg-type]
            output_path,
        )


def test_save_overwrites_existing_file_when_enabled(
    fake_report: FakePortfolioReport,
    tmp_path: Path,
) -> None:
    output_path = (
        tmp_path
        / "portfolio.xlsx"
    )

    output_path.write_bytes(
        b"existing"
    )

    service = ExcelReportService()

    result = service.save(
        fake_report,  # type: ignore[arg-type]
        output_path,
        overwrite=True,
    )

    assert result.read_bytes().startswith(
        b"PK"
    )


@pytest.mark.parametrize(
    ("parameter_name", "kwargs"),
    [
        (
            "create_parent_directories",
            {
                "create_parent_directories": 1,
            },
        ),
        (
            "overwrite",
            {
                "overwrite": 1,
            },
        ),
    ],
)
def test_save_rejects_invalid_boolean_options(
    fake_report: FakePortfolioReport,
    tmp_path: Path,
    parameter_name: str,
    kwargs: dict[str, object],
) -> None:
    service = ExcelReportService()

    with pytest.raises(
        TypeError,
        match=f"{parameter_name} must be a boolean",
    ):
        service.save(
            fake_report,  # type: ignore[arg-type]
            tmp_path / "portfolio.xlsx",
            **kwargs,  # type: ignore[arg-type]
        )


def test_save_wraps_file_write_failure(
    fake_report: FakePortfolioReport,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    service = ExcelReportService()

    monkeypatch.setattr(
        Path,
        "write_bytes",
        Mock(
            side_effect=OSError(
                "simulated write failure"
            )
        ),
    )

    with pytest.raises(
        ExcelReportGenerationError,
        match=(
            "Unable to save the portfolio Excel report: "
            "simulated write failure"
        ),
    ) as error_info:
        service.save(
            fake_report,  # type: ignore[arg-type]
            tmp_path / "portfolio.xlsx",
        )

    assert isinstance(
        error_info.value.__cause__,
        OSError,
    )


# ============================================================
# Download Preparation
# ============================================================


def test_prepare_download_returns_expected_payload(
    fake_report: FakePortfolioReport,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    expected_bytes = b"PK-test-workbook"

    generate_mock = Mock(
        return_value=expected_bytes
    )

    monkeypatch.setattr(
        ExcelReportService,
        "generate_bytes",
        generate_mock,
    )

    service = ExcelReportService()

    result = service.prepare_download(
        fake_report,  # type: ignore[arg-type]
        filename="portfolio-analysis.xlsx",
    )

    assert result == (
        expected_bytes,
        "portfolio-analysis.xlsx",
        EXCEL_MIME_TYPE,
    )

    generate_mock.assert_called_once_with(
        fake_report
    )


def test_prepare_download_uses_default_filename(
    fake_report: FakePortfolioReport,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(
        ExcelReportService,
        "generate_bytes",
        lambda self, report: b"PK-workbook",
    )

    service = ExcelReportService()

    workbook_bytes, filename, mime_type = (
        service.prepare_download(
            fake_report  # type: ignore[arg-type]
        )
    )

    assert workbook_bytes == b"PK-workbook"
    assert filename == DEFAULT_EXCEL_FILENAME
    assert mime_type == EXCEL_MIME_TYPE


@pytest.mark.parametrize(
    "filename",
    [
        "",
        " ",
        "portfolio.xls",
        "portfolio.pdf",
        "folder/portfolio.xlsx",
        "../portfolio.xlsx",
    ],
)
def test_prepare_download_rejects_invalid_filename_before_generation(
    fake_report: FakePortfolioReport,
    filename: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    generate_mock = Mock()

    monkeypatch.setattr(
        ExcelReportService,
        "generate_bytes",
        generate_mock,
    )

    service = ExcelReportService()

    with pytest.raises(
        ExcelReportValidationError,
    ):
        service.prepare_download(
            fake_report,  # type: ignore[arg-type]
            filename=filename,
        )

    generate_mock.assert_not_called()


# ============================================================
# Convenience APIs
# ============================================================


def test_generate_portfolio_excel_delegates_to_service(
    fake_report: FakePortfolioReport,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    expected_bytes = b"PK-convenience"

    generate_mock = Mock(
        return_value=expected_bytes
    )

    monkeypatch.setattr(
        ExcelReportService,
        "generate_bytes",
        generate_mock,
    )

    result = generate_portfolio_excel(
        fake_report  # type: ignore[arg-type]
    )

    assert result == expected_bytes

    generate_mock.assert_called_once_with(
        fake_report
    )


def test_save_portfolio_excel_delegates_to_service(
    fake_report: FakePortfolioReport,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    expected_path = (
        tmp_path
        / "portfolio.xlsx"
    ).resolve()

    save_mock = Mock(
        return_value=expected_path
    )

    monkeypatch.setattr(
        ExcelReportService,
        "save",
        save_mock,
    )

    result = save_portfolio_excel(
        fake_report,  # type: ignore[arg-type]
        expected_path,
        create_parent_directories=False,
        overwrite=True,
    )

    assert result == expected_path

    save_mock.assert_called_once_with(
        fake_report,
        expected_path,
        create_parent_directories=False,
        overwrite=True,
    )


def test_generate_portfolio_excel_propagates_validation_error() -> None:
    with pytest.raises(
        TypeError,
        match="report must be a PortfolioReport",
    ):
        generate_portfolio_excel(
            None  # type: ignore[arg-type]
        )