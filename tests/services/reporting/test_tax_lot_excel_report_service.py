"""Tests for tax-lot Excel report generation."""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from services.reporting.tax_lot_excel_report_service import (
    TaxLotExcelReportService,
)
from services.tax_lot_service import (
    RealizedGain,
    TaxLot,
    TaxLotAnalysis,
)


def test_generate_bytes_builds_open_tax_lots_sheet() -> None:
    """Open tax lots should be exported as a structured worksheet."""

    analysis = TaxLotAnalysis(
        open_lots=[
            TaxLot(
                portfolio_id=10,
                fund_id=20,
                source_transaction_id=1,
                acquisition_date=date(2026, 1, 10),
                original_units=Decimal("100.000000"),
                remaining_units=Decimal("60.000000"),
                cost_per_unit=Decimal("10.0000"),
                remaining_cost_basis=Decimal("600.00"),
            )
        ],
        realized_gains=[],
    )

    excel_bytes = TaxLotExcelReportService().generate_bytes(
        analysis
    )

    workbook = load_workbook(
        BytesIO(excel_bytes),
    )

    try:
        assert "Open Tax Lots" in workbook.sheetnames

        worksheet = workbook["Open Tax Lots"]

        assert [
            worksheet.cell(row=3, column=column).value
            for column in range(1, 9)
        ] == [
            "Portfolio ID",
            "Fund ID",
            "Source Transaction ID",
            "Acquisition Date",
            "Original Units",
            "Remaining Units",
            "Cost Per Unit",
            "Remaining Cost Basis",
        ]

        assert worksheet["A4"].value == 10
        assert worksheet["B4"].value == 20
        assert worksheet["C4"].value == 1
        assert worksheet["D4"].value == datetime(
            2026,
            1,
            10,
            0,
            0,
        )
        assert worksheet["E4"].value == 100.0
        assert worksheet["F4"].value == 60.0
        assert worksheet["G4"].value == 10.0
        assert worksheet["H4"].value == 600.0

        assert worksheet.freeze_panes == "A4"
        assert worksheet.auto_filter.ref == "A3:H4"

    finally:
        workbook.close()
def test_generate_bytes_builds_realized_gains_sheet() -> None:
    """Realized FIFO matches should be exported to a worksheet."""

    analysis = TaxLotAnalysis(
        open_lots=[],
        realized_gains=[
            RealizedGain(
                portfolio_id=10,
                fund_id=20,
                sell_transaction_id=2,
                source_buy_transaction_id=1,
                acquisition_date=date(2026, 1, 10),
                disposal_date=date(2026, 2, 10),
                holding_period_days=31,
                units_sold=Decimal("40.000000"),
                sale_proceeds=Decimal("600.00"),
                cost_basis=Decimal("400.00"),
                realized_gain=Decimal("200.00"),
            )
        ],
    )

    excel_bytes = TaxLotExcelReportService().generate_bytes(
        analysis
    )

    workbook = load_workbook(
        BytesIO(excel_bytes),
    )

    try:
        assert "Realized Gains" in workbook.sheetnames

        worksheet = workbook["Realized Gains"]

        assert [
            worksheet.cell(row=3, column=column).value
            for column in range(1, 12)
        ] == [
            "Portfolio ID",
            "Fund ID",
            "Sell Transaction ID",
            "Source Buy Transaction ID",
            "Acquisition Date",
            "Disposal Date",
            "Holding Period Days",
            "Units Sold",
            "Sale Proceeds",
            "Cost Basis",
            "Realized Gain",
        ]

        assert worksheet["A4"].value == 10
        assert worksheet["B4"].value == 20
        assert worksheet["C4"].value == 2
        assert worksheet["D4"].value == 1
        assert worksheet["E4"].value == datetime(
            2026,
            1,
            10,
            0,
            0,
        )
        assert worksheet["F4"].value == datetime(
            2026,
            2,
            10,
            0,
            0,
        )
        assert worksheet["G4"].value == 31
        assert worksheet["H4"].value == 40.0
        assert worksheet["I4"].value == 600.0
        assert worksheet["J4"].value == 400.0
        assert worksheet["K4"].value == 200.0

        assert worksheet.freeze_panes == "A4"
        assert worksheet.auto_filter.ref == "A3:K4"

    finally:
        workbook.close()
def test_prepare_download_returns_bytes_and_metadata() -> None:
    """Download preparation should return XLSX data and metadata."""

    analysis = TaxLotAnalysis(
        open_lots=[],
        realized_gains=[],
    )

    service = TaxLotExcelReportService()

    excel_bytes, filename, mime_type = (
        service.prepare_download(
            analysis
        )
    )

    assert excel_bytes.startswith(b"PK")
    assert filename == "portfolio_tax_lots.xlsx"
    assert mime_type == (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )